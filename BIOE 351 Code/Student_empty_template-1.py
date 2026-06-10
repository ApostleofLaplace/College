# -*- coding: utf-8 -*-
"""
Created on Wed May  6 13:19:30 2026

@author: prickets
"""

'''
This code workshop is to teach basic coding skills for data analysis as it pertains to contact angles for Biomaterials BIOE 351 Spring 2026. To be given in weeks 3-7 Week 4 maybe. (MAY 13th)

# ============================================================================
# PART 5: To do as homework - work together before you ask AI.
# ============================================================================
#A: Find PE surface energy

# Now it is your turn to use the polar and dispersive components to find the surface energy for polyehtylene. Don't hard code the values, use python code to find the values in the excel file, apply the necessary calculation, and have it print into the console. 

# If you are really good, you'll be able to use code to update your saved copy of the template in the correct excel cell.



# B: Find the theta for each solvent on clean glass and update the excel file.



# ============================================================================
# PART 6: ZISMAN PLOT FOR POLYETHYLENE - given- Make a new one for PTFE.
# ============================================================================



'''

#import modules - what are modules?
import os
import pandas as pd
#   https://pandas.pydata.org/docs/user_guide/dsintro.html
import matplotlib as plt
import numpy as np
from scipy.optimize import fsolve
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font

# ============================================================================
# Grab excel file for manipulation and data overwrite
# ============================================================================
xlsx_path = 


# ============================================================================
# PART 2: THEORY - THE OWRK EQUATION
# ============================================================================
"""
OWRK (Owens-Wendt-Rabel-Kaelble) Equation:
Y_l * (1 + cos(theta)) = 2 * sqrt(Y_ds * Y_dl) + 2 * sqrt(Y_ps * Y_pl)
 
Where:
- Y_l = liquid surface tension (measured with Wilhelmy plate)
- theta = contact angle on a surface
- Y_ds, Y_ps = dispersive and polar components of the SOLID surface energy
- Y_dl, Y_pl = dispersive and polar components of the LIQUID
- Y_total = Y_dispersive + Y_polar
 
Key Insight:
PTFE has Y_p ≈ 0 (no polar component), so we can use it to find Y_dl!

# Y_l = Y_p + Y_d




# Equations with defined variables:
# equation_1:
# (Y_l * (1 + cos(theta_PTFE)) = 2 * sqrt(Y_d_ptfe * Y_dl)) = 0  because Y_p_ptfe for water on ptfe = 0.

# equation_2:
# Y_l * (1 + cos(theta_glass)) = 2 * sqrt(Y_d_glass * Y_dl)) + 2 * sqrt(Y_d_glass * Y_dl)) 

# Use equation_1 to solve for Y_dl, then plug that into equation_2.


"""



# ============================================================================
# PART 3: CALCULATE POLAR AND DISPERSIVE COMPONENTS OF LIQUIDS
# ============================================================================
 
# Known values for PTFE from the excel file, how to call?
#



#Use Scipy to solve equation 1
# 1. Known parameters (example values) iloc starts searching pythonically after the specified header row.







#Now it is your turn, try to write some code so it takes the result from solving equation_1 and plugs it into solving equation_2!


print(f"\n\n\n")

#Now, we could enter this into our excel sheet by hand, but let's see if we can get through this workshop and get our code to do it at the end.




# ============================================================================
#  Use a LOOP and iterate over all values in the column!
# ============================================================================

#Trying to code for each polar and dispersive component will be time-consuming. Let's make a loop to iteratively calculate each solvent's components automagically.


# Empty lists:




#Build the loop, note that this takes case-sensitive strings that match exactly to the excel file.

    
    #calculate for THIS liquid:
    theta_rad = np.radians(theta_ptfe) #convert from degrees to radians so the cos() takes it correctly.
    
    #re-write function to become a generalized one.
    
    
    try:
        #Run each one through the equation solver function we wrote, INTRODUCE TRY, EXCEPT, RAISE ERROR BLOCKS
          # Initial guess = 30% of total
         # Keep in bounds: min(Y_dl,Ydl) returns Y_dl itself, compares Y_dl to 0 (max(0...), IF Y_dl is (-), then returns 0 for Y_dl.
        
        
        
        
    
    except:
       
        #may get run time error.
    
    
    
    





# ============================================================================
# PART 4: COPY TEMPLATE AND FILL IN CALCULATED VALUES
# ============================================================================

print("\n" + "="*80)
print("SAVING YOUR RESULTS")
print("="*80)

# Step 1: Create output filename
student_name_clean = student_name.replace(" ", "_")
output_filename = f"contact_angle_RESULTS_{student_name_clean}.xlsx"
output_path = os.path.join(folder_path, output_filename)

# Step 2: Copy the original file (template)
shutil.copy(xlsx_path, output_path)
print(f"✓ Copied template to: {output_path}")

# Step 3: Open the copied file with openpyxl
wb = load_workbook(output_path)
ws = wb.active  # or wb['Sheet1'] if you know the sheet name

# Step 4: Find which columns are 'dispersive' and 'polar'
# Look at header row (row 3 since header=2 means Excel row 3)
header_row = 3
dispersive_col = None
polar_col = None

for col_idx, cell in enumerate(ws[header_row], start=1):
    if cell.value == 'dispersive_a':
        dispersive_col = col_idx
    elif cell.value == 'polar_a':
        polar_col = col_idx

print(f"✓ Found dispersive column: {dispersive_col}")
print(f"✓ Found polar column: {polar_col}")

# Step 5: Fill in the calculated values
# Data starts at row 4 (row 3 is headers)
data_start_row = 4

for i, (disp_val, polar_val) in enumerate(zip(dispersive_components, polar_components)):
    row_num = data_start_row + i
    
    # Write dispersive value
    if dispersive_col:
        ws.cell(row=row_num, column=dispersive_col, value=disp_val)
    
    # Write polar value
    if polar_col:
        ws.cell(row=row_num, column=polar_col, value=polar_val)

print(f"✓ Filled in {len(dispersive_components)} rows of data")

# Step 6: Add student name to row 1
ws['A1'] = f"Analyzed by: {student_name}"
ws['A1'].font = Font(bold=True, size=14)

# Step 7: Save the workbook
wb.save(output_path)
wb.close()

print(f"✓ File saved successfully!")
print(f"✓ Original template unchanged")
print(f"✓ Your file: {output_path}")

# ============================================================================
# PART 5: To do as homework - work together before you ask AI.
# ============================================================================
#A: Find PE surface energy

# Now it is your turn to use the polar and dispersive components to find the surface energy for polyehtylene. Don't hard code the values, use python code to find the values in the excel file, apply the necessary calculation, and have it print into the console. 

# If you are really good, you'll be able to use code to update your saved copy of the template in the correct excel cell.



# B: Find the theta for each solvent on clean glass and update the excel file.



# ============================================================================
# PART 6: ZISMAN PLOT FOR POLYETHYLENE - given- Make a new one for PTFE.
# ============================================================================


import matplotlib.pyplot as plt
from scipy.stats import linregress

print("\n" + "="*80)
print("ZISMAN PLOT FOR POLYETHYLENE")
print("="*80)

# Step 1: Collect data for liquids with PE contact angles
print("\nCollecting data for polyethylene...")

liquids_pe = []
surface_tensions = []
cos_thetas = []

for idx, row in idat.iterrows():
    solvent = row['Solvent']
    theta_pe = row['theta on clean Polyethylene']
    Y_l = row['Surface tension Dyne/cm = miliN/meter']
    
    # Skip if missing data
    if pd.isna(theta_pe) or pd.isna(Y_l):
        continue
    
    # Calculate cos(theta)
    theta_rad = np.radians(theta_pe)
    cos_theta = np.cos(theta_rad)
    
    liquids_pe.append(solvent)
    surface_tensions.append(Y_l)
    cos_thetas.append(cos_theta)
    
    print(f"  {solvent:<20} Y_l={Y_l:.1f} mN/m, θ={theta_pe:.1f}°, cos(θ)={cos_theta:.4f}")

print(f"\n✓ Found {len(liquids_pe)} liquids with PE data")

# Step 2: Linear regression
if len(liquids_pe) >= 2:
    slope, intercept, r_value, p_value, std_err = linregress(surface_tensions, cos_thetas)
    
    print(f"\nLinear fit: cos(θ) = {slope:.6f} * γ_l + {intercept:.4f}")
    print(f"R² = {r_value**2:.4f}")
    
    # Step 3: Calculate critical surface tension (where cos(θ) = 1)
    # cos(θ) = 1 when θ = 0° (complete wetting)
    # 1 = slope * γ_c + intercept
    # γ_c = (1 - intercept) / slope
    
    gamma_critical = (1 - intercept) / slope
    
    print(f"\n✓ CRITICAL SURFACE TENSION of Polyethylene:")
    print(f"  γ_c = {gamma_critical:.2f} mN/m")
    print(f"  (Literature value for PE: ~31-33 mN/m)")
    
    # Step 4: Create Zisman plot
    fig, ax = plt.subplots(figsize=(10, 7))
    
    # Plot data points
    ax.scatter(surface_tensions, cos_thetas, s=100, c='steelblue', 
               edgecolors='black', linewidth=1.5, zorder=3, label='Measured')
    
    # Add labels for each point
    for i, solvent in enumerate(liquids_pe):
        ax.annotate(solvent.strip(), 
                   (surface_tensions[i], cos_thetas[i]),
                   xytext=(5, 5), textcoords='offset points',
                   fontsize=9, alpha=0.8)
    
    # Plot regression line
    x_fit = np.linspace(min(surface_tensions)*0.9, max(surface_tensions)*1.1, 100)
    y_fit = slope * x_fit + intercept
    ax.plot(x_fit, y_fit, 'r--', linewidth=2, label=f'Linear fit (R²={r_value**2:.3f})')
    
    # Extend line to critical surface tension
    x_extended = np.linspace(min(surface_tensions)*0.9, gamma_critical*1.1, 100)
    y_extended = slope * x_extended + intercept
    ax.plot(x_extended, y_extended, 'r:', linewidth=1.5, alpha=0.5)
    
    # Mark critical surface tension
    ax.axhline(y=1.0, color='green', linestyle='--', linewidth=1.5, alpha=0.7, label='cos(θ) = 1 (complete wetting)')
    ax.axvline(x=gamma_critical, color='orange', linestyle='--', linewidth=1.5, alpha=0.7)
    ax.plot(gamma_critical, 1.0, 'o', color='red', markersize=12, zorder=4, 
            label=f'γ_c = {gamma_critical:.1f} mN/m')
    
    # Formatting
    ax.set_xlabel('Liquid Surface Tension, γ_l (mN/m)', fontsize=12, fontweight='bold')
    ax.set_ylabel('cos(θ)', fontsize=12, fontweight='bold')
    ax.set_title('Zisman Plot for Polyethylene\nCritical Surface Tension Determination', 
                fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3)
    ax.legend(loc='best', fontsize=10)
    
    # Set y-axis limits to show complete range
    ax.set_ylim([min(cos_thetas)*0.95, 1.05])
    ax.set_xlim([min(surface_tensions)*0.9, max(max(surface_tensions), gamma_critical)*1.1])
    
    # Add equation text box
    textstr = f'cos(θ) = {slope:.4f}γ_l + {intercept:.3f}\nR² = {r_value**2:.4f}\nγ_c = {gamma_critical:.2f} mN/m'
    props = dict(boxstyle='round', facecolor='wheat', alpha=0.8)
    ax.text(0.05, 0.05, textstr, transform=ax.transAxes, fontsize=11,
            verticalalignment='bottom', bbox=props)
    
    plt.tight_layout()
    
    # Save plot
    plot_path = os.path.join(folder_path, f'Zisman_plot_{student_name}.png')
    plt.savefig(plot_path, dpi=300, bbox_inches='tight')
    print(f"\n✓ Zisman plot saved to: {plot_path}")
    plt.show()

else:
    print("\n⚠ Not enough data points for Zisman plot (need at least 2)")

print("\n" + "="*80)
print("DISCUSSION QUESTIONS:")
print("="*80)
print("""
1. What does the critical surface tension tell us about polyethylene?
  

2. How does PE's critical surface tension (~31-33 mN/m) compare to:
   → PTFE: ~18 mN/m (more hydrophobic)
   → Glass: ~60+ mN/m (more hydrophilic)

3. Why is the Zisman plot useful?
 

4. What does the R² value tell you?
  
""")


