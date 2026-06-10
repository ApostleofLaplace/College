# Modules

import os
import pandas as pd
import matplotlib as plt
import numpy as np
from scipy.optimize import fsolve
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font

# File path stuff

xcel_dir = os.path.dirname(os.path.abspath(__file__))
xcel_file = "TEMPLATE_contact_angle_BIOE_351.xlsx"

xcel_path = os.path.join(xcel_dir, xcel_file)

xcel_data = pd.read_excel(xcel_path, sheet_name=0, header=2, engine='openpyxl')

student_name = 'Trey_Cherry'

# Hardcoded variables

y_s_ptfe = 18
y_p_ptfe = 0
y_d_ptfe = 18

# Shit im ripping from the xlsx

y_l_water = xcel_data.iloc[0,1] 

theta_ptfe_deg = xcel_data.iloc[0,2]
theta_ptfe_rad = np.radians(theta_ptfe_deg)

def goofy_silly_number_diddler(y_dl):
    
    lhs = y_l_water * (1 + np.cos(theta_ptfe_rad))
    rhs = 2 * np.sqrt(y_d_ptfe * y_dl)
    
    dif_left_right = lhs - rhs
    
    return dif_left_right 

y_dl_guess = 1.0 

y_dl_1_water = fsolve(goofy_silly_number_diddler, y_dl_guess)

print(y_dl_1_water)

dispersive_components = []
polar_components = []

for idx, row in xcel_data.iterrows():
    
    solvent = row['Solvent']
    y_l = row['tension']
    
    theta_ptfe = row['theta on PTFE']
    theta_rad = np.radians(theta_ptfe)
    
    def equation_1(y_dl):
        
        lhs = y_l * (1 + np.cos(theta_rad))
        rhs = 2 * np.sqrt(y_d_ptfe * y_dl)
        
        return lhs - rhs
    
    try: 
        
        y_dl = fsolve(equation_1, y_l*0.3)[0]  # Extract scalar from array
        y_dl = max(0, min(y_dl, y_l))
        
        y_pl = y_l - y_dl 
        
        dispersive_components.append(y_dl)
        polar_components.append(y_pl)
        
        print(f'{solvent}: y_d = {y_dl}, y_p = {y_pl}')
        
        
    except:
        
        dispersive_components.append(np.nan)
        polar_components.append(np.nan)
        
        raise ValueError('it didnt work dumbass')

output_filename = f"{student_name}_contact_angle_results.xlsx"
output_path = os.path.join(xcel_dir, output_filename)

shutil.copy(xcel_path, output_path)

wb = load_workbook(output_path)
ws = wb.active

ws['A1'] = f"Student: {student_name}"
ws['A1'].font = Font(bold=True, size=12)

header_row = 3
dispersive_col = None
polar_col = None

for col_idx, cell in enumerate(ws.iter_cols(min_row=header_row, max_row=header_row), 1):
    cell_value = cell[0].value
    if cell_value == 'dispersive_a':
        dispersive_col = col_idx
    elif cell_value == 'polar_a':
        polar_col = col_idx

if dispersive_col is None or polar_col is None:
    raise ValueError(f"Could not find columns: dispersive_a (col {dispersive_col}), polar_a (col {polar_col})")

start_row = 4
for idx, (y_dl, y_pl) in enumerate(zip(dispersive_components, polar_components)):
    row_num = start_row + idx
    ws.cell(row=row_num, column=dispersive_col, value=float(y_dl))
    ws.cell(row=row_num, column=polar_col, value=float(y_pl) if not np.isnan(y_pl) else None)

wb.save(output_path)
print(f"\n✓ Results saved to: {output_path}")

glass_row_idx = None
for idx, row in xcel_data.iterrows():
    if row.get('material', '').lower() == 'glass':
        glass_row_idx = idx
        break

if glass_row_idx is not None:
    y_d_glass_from_file = xcel_data.iloc[glass_row_idx].get('dispersive_a')
    y_p_glass_from_file = xcel_data.iloc[glass_row_idx].get('polar_a')
    
    y_s_pe = None
    y_p_pe = None
    y_d_pe = None
    
    print("\n=== PART 5: PE Surface Energy ===")
    print("Note: Requires glass contact angle data from Excel")
    print(f"PE Surface Energy (if calculated): {y_s_pe} mN/m")
else:
    print("\nPart 5: Glass data not found")


surface_tensions = []
cos_theta_pe = []
solvent_names = []

for idx, row in xcel_data.iterrows():
    y_l = row.get('tension')
    theta_pe = row.get('theta on clean Polyethylene')
    solvent = row.get('Solvent')
    
    if pd.notna(y_l) and pd.notna(theta_pe):
        surface_tensions.append(y_l)
        cos_theta_pe.append(np.cos(np.radians(theta_pe)))
        solvent_names.append(solvent)

if surface_tensions:
   
    coefficients = np.polyfit(surface_tensions, cos_theta_pe, 1)
    slope = coefficients[0]
    intercept = coefficients[1]

    if slope != 0:
        gamma_c = (1 - intercept) / slope
        print(f"Critical Surface Tension (γc) for PE: {gamma_c:.2f} mN/m")

    import matplotlib.pyplot as plt

    plt.figure(figsize=(10, 6))
    plt.plot(surface_tensions, cos_theta_pe, 'bo-', label='PE data')
    x_fit = np.linspace(min(surface_tensions)-5, max(surface_tensions)+5, 100)
    y_fit = slope * x_fit + intercept
    plt.plot(x_fit, y_fit, 'r--', label='Linear fit')
    plt.axhline(y=1, color='k', linestyle=':', alpha=0.5)
    plt.axvline(x=gamma_c, color='g', linestyle=':', alpha=0.5, label=f'γc = {gamma_c:.2f} mN/m')
    plt.xlabel('Surface Tension of Liquid (mN/m)')
    plt.ylabel('cos(θ) on Polyethylene')
    plt.title('Zisman Plot for Polyethylene')
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.show()


else:
    print("No polyethylene contact angle data found for Zisman plot")